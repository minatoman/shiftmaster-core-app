# ---- Django標準 ----
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.forms import modelformset_factory
from django.core.exceptions import PermissionDenied, ValidationError

# ---- 外部ライブラリ ----
from xhtml2pdf import pisa
from io import BytesIO, StringIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import calendar

# ---- Python標準ライブラリ ----
import os
import io
import json
import csv
import tempfile
from datetime import date, datetime
from pathlib import Path

# ---- 自作ユーティリティ ----
from .utils import normalize_name, process_uploaded_files  # 修正：必要なユーティリティのインポート

# ---- アプリ内モデル／フォーム ----
from .models import (
    Shift, Employee, ShiftRequest, HolidayRequest,
    LeaveBalance, Committee, OnCallShift, ShiftRequirement,
    DialysisPatient, DialysisStaff
)
from .forms import (
    ShiftForm, ShiftRequestForm, ShiftApprovalForm,
    HolidayRequestForm, ShiftEditForm, DialysisPatientForm, DialysisStaffForm,
    EmployeeForm, UploadFileForm  # 修正：UploadFileFormのインポートを追加
)

# ✅ 権限チェック
def staff_required(view_func):
    return user_passes_test(lambda u: u.is_staff)(view_func)

# ✅ メール通知関数
def send_notification_email(subject, message, recipient_email): 
    try:
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient_email])
    except Exception as e:
        print(f"メール送信失敗: {e}")

# 📅 カレンダービュー
@login_required
def calendar_view(request):
    return render(request, 'shifts/calendar_full.html')

@login_required
def calendar_event_api(request):
    events = []
    for shift in Shift.objects.all():
        events.append({
            "title": f"[確定] {shift.employee.name}（{shift.shift_type}）",
            "start": str(shift.shift_date),
            "color": "#007bff"
        })
    for req in ShiftRequest.objects.all():
        events.append({
            "title": f"[希望] {req.employee.name}（{req.shift_type}）",
            "start": str(req.requested_date),
            "color": "#28a745"
        })
    return JsonResponse(events, safe=False)

# ✅ 透析日誌 関連ビュー
@login_required
def dialysis_daily_view(request, selected_date=None):
    selected_date = selected_date or date.today()
    patients = DialysisPatient.objects.filter(date=selected_date)
    staff = DialysisStaff.objects.filter(date=selected_date)
    return render(request, 'dialysis/daily_view.html', {
        'patients': patients, 'staff': staff, 'date': selected_date
    })

@login_required
def dialysis_register_view(request):
    if not request.user.is_staff and datetime.now().day > 17:
        raise PermissionDenied("17日以降は編集できません（スタッフ専用）")

    PatientFormSet = modelformset_factory(DialysisPatient, form=DialysisPatientForm, extra=5)
    StaffFormSet = modelformset_factory(DialysisStaff, form=DialysisStaffForm, extra=5)

    if request.method == 'POST':
        pf = PatientFormSet(request.POST, queryset=DialysisPatient.objects.none())
        sf = StaffFormSet(request.POST, queryset=DialysisStaff.objects.none())
        if pf.is_valid() and sf.is_valid():
            pf.save()
            sf.save()
            return redirect('dialysis_daily')
    else:
        pf = PatientFormSet(queryset=DialysisPatient.objects.none())
        sf = StaffFormSet(queryset=DialysisStaff.objects.none())

    return render(request, 'dialysis/register.html', {'pf': pf, 'sf': sf})

@login_required
def import_dialysis_csv(request):
    if request.method == "POST" and request.FILES.get('file'):
        df = pd.read_excel(request.FILES['file'])
        for _, row in df.iterrows():
            if not pd.isna(row['名前']):
                DialysisPatient.objects.create(
                    name=row['名前'],
                    dialysis_time=row['透析時間'],
                    gender=row['性別'],
                    area=row['地域'],
                    remarks=row.get('備考', ''),
                    date=row.get('日付', date.today())
                )
        messages.success(request, "透析日誌データをインポートしました。")
        return redirect('dialysis_daily')
    return render(request, 'dialysis/import.html')

@login_required
def dialysis_pdf_export(request, date_str):
    target_date = date.fromisoformat(date_str)
    patients = DialysisPatient.objects.filter(date=target_date)
    staff = DialysisStaff.objects.filter(date=target_date)
    template = get_template('dialysis/pdf_template.html')
    html = template.render({'patients': patients, 'staff': staff, 'date': target_date})
    response = BytesIO()
    pisa.CreatePDF(html, dest=response)
    return HttpResponse(response.getvalue(), content_type='application/pdf')

@login_required
def dialysis_calendar_api_events(request):
    events = []
    dates = DialysisPatient.objects.values_list('date', flat=True).distinct()
    for d in dates:
        events.append({'title': '透析あり', 'start': d.isoformat()})
    return JsonResponse(events, safe=False)

# 🧑‍💼 ユーザー登録／ログイン／ログアウト
def signup_view(request):
    form = UserCreationForm(request.POST or None)
    if form.is_valid():
        user = form.save()
        messages.success(request, 'ユーザー登録完了。ログインしてください。')
        admin_emails = [emp.email for emp in Employee.objects.filter(position="管理者")]
        for email in admin_emails:
            send_notification_email(
                subject="【新規ユーザー登録】",
                message=f"新しいユーザー {user.username} さんが登録されました。",
                recipient_email=email
            )
        return redirect('login')
    elif request.method == 'POST':
        messages.error(request, '登録に失敗しました。入力内容をご確認ください。')
    return render(request, 'registration/signup.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def profile(request):
    return render(request, 'shifts/profile.html', {'user': request.user})

@login_required
def homepage(request):
    return render(request, 'shifts/homepage.html')

# 📥 勤務希望（JS API登録＋17日制限）
@csrf_exempt
@login_required
def shift_request_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("date")
            shift_type = data.get("type", "希望")
            requested_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            if requested_date.day > 17 and not request.user.is_staff:
                return JsonResponse({"error": "17日以降の希望登録はできません"}, status=400)

            ShiftRequest.objects.create(
                employee=request.user.employee,
                requested_date=requested_date,
                shift_type=shift_type,
                priority=1
            )
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

# 📋 勤務希望一覧／休暇申請
@login_required
def shift_request_list(request):
    requests = ShiftRequest.objects.all().order_by('requested_date')
    return render(request, 'shifts/shift_request_list.html', {'requests': requests})

@login_required
def add_holiday_request(request):
    form = HolidayRequestForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, '休暇申請が完了しました！')
        return redirect('shift_request_list')
    return render(request, 'shifts/add_holiday_request.html', {'form': form})

# 📤 勤務希望出力（Excel / CSV / PDF）
@login_required
def export_shift_requests_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務希望"
    ws.append(["氏名", "希望日", "勤務区分", "優先度"])
    for r in ShiftRequest.objects.all():
        ws.append([r.employee.name, r.requested_date, r.shift_type, r.priority])

    output_path = os.path.join(settings.BASE_DIR, 'shifts', 'temp_exports', 'shift_requests.xlsx')
    wb.save(output_path)

    with open(output_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename="shift_requests.xlsx"'
    return response

@login_required
def export_shift_requests_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="shift_requests.csv"'
    writer = csv.writer(response)
    writer.writerow(['氏名', '希望日', '勤務区分', '優先度'])
    for r in ShiftRequest.objects.all():
        writer.writerow([r.employee.name, r.requested_date, r.shift_type, r.priority])
    return response

@login_required
def export_shift_pdf(request):
    shifts = Shift.objects.order_by('shift_date')
    today = datetime.now()
    html = get_template('shifts/shift_pdf.html').render({'shifts': shifts, 'today': today})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="shift_report.pdf"'
    pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    return response

# 📥 インポート（Excel）
@login_required
def import_shift_requests_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        wb = load_workbook(request.FILES['excel_file'])
        ws = wb.active
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            try:
                name = normalize_name(row[0])
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row[1],
                    shift_type=row[2],
                    priority=row[3] or 1
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row[0]}")
            except Exception as e:
                messages.warning(request, f"エラー（{row[0]}）：{e}")
        messages.success(request, 'Excelインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_excel.html')

# 📥 インポート（CSV）
@login_required
def import_shift_requests_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        df = pd.read_csv(request.FILES['csv_file'])
        for _, row in df.iterrows():
            try:
                name = normalize_name(row.get('氏名', ''))
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row['希望日'],
                    shift_type=row['勤務区分'],
                    priority=row.get('優先度', 1)
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row.get('氏名')}")
            except Exception as e:
                messages.warning(request, f"エラー（{row.get('氏名')}）：{e}")
        messages.success(request, 'CSVインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_csv.html')

# 🤖 自動割当（ルールベースAI風）
@staff_required
def auto_assign(request):
    assigned = set()
    assigned_counts = defaultdict(lambda: defaultdict(int))
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
        week_day_index = list(calendar.day_name).index(day) + 2
        requests = ShiftRequest.objects.filter(requested_date__week_day=week_day_index).order_by('priority')
        requirements = ShiftRequirement.objects.filter(day_of_week=day).first()
        if not requirements:
            continue
        counts = {
            '看護師': requirements.nurse_required,
            '臨床工学技士': requirements.engineer_required,
            '介護福祉士': requirements.assistant_required
        }
        for req in requests:
            emp_id = req.employee.id
            date_key = (emp_id, req.requested_date)
            position = req.employee.position
            if counts.get(position, 0) > 0 and date_key not in assigned:
                Shift.objects.create(
                    employee=req.employee,
                    shift_date=req.requested_date,
                    shift_type=req.shift_type,
                    is_approved=True
                )
                assigned.add(date_key)
                assigned_counts[emp_id][day] += 1
                counts[position] -= 1
    messages.success(request, '最適化された勤務割当が完了しました。')
    return redirect('shift_list')
