# shifts/views/request_views.py

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from datetime import datetime
import json
import os
import csv
import pandas as pd
from openpyxl import load_workbook, Workbook

from django.conf import settings
from shifts.models import ShiftRequest, Employee
from shifts.forms import HolidayRequestForm
from shifts.utils import normalize_name


@csrf_exempt
@login_required
def shift_request_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("date")
            shift_type = data.get("type", "希望")
            requested_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            if requested_date.day > 17 and not request.user.is_staff:
                return JsonResponse({"error": "17日以降の希望登録はできません"}, status=400)

            ShiftRequest.objects.create(
                employee=request.user.employee,
                requested_date=requested_date,
                shift_type=shift_type,
                priority=1
            )
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


@login_required
def shift_request_list(request):
    requests = ShiftRequest.objects.all().order_by('requested_date')
    return render(request, 'shifts/shift_request_list.html', {'requests': requests})


@login_required
def add_holiday_request(request):
    form = HolidayRequestForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, '休暇申請が完了しました！')
        return redirect('shift_request_list')
    return render(request, 'shifts/add_holiday_request.html', {'form': form})


@login_required
def export_shift_requests_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務希望"
    ws.append(["氏名", "希望日", "勤務区分", "優先度"])
    for r in ShiftRequest.objects.all():
        ws.append([r.employee.name, r.requested_date, r.shift_type, r.priority])

    output_path = os.path.join(settings.BASE_DIR, 'shifts', 'temp_exports', 'shift_requests.xlsx')
    wb.save(output_path)

    with open(output_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename="shift_requests.xlsx"'
    return response


@login_required
def export_shift_requests_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="shift_requests.csv"'
    writer = csv.writer(response)
    writer.writerow(['氏名', '希望日', '勤務区分', '優先度'])
    for r in ShiftRequest.objects.all():
        writer.writerow([r.employee.name, r.requested_date, r.shift_type, r.priority])
    return response


@login_required
def import_shift_requests_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        wb = load_workbook(request.FILES['excel_file'])
        ws = wb.active
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            try:
                name = normalize_name(row[0])
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row[1],
                    shift_type=row[2],
                    priority=row[3] or 1
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row[0]}")
            except Exception as e:
                messages.warning(request, f"エラー（{row[0]}）：{e}")
        messages.success(request, 'Excelインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_excel.html')


@login_required
def import_shift_requests_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        df = pd.read_csv(request.FILES['csv_file'])
        for _, row in df.iterrows():
            try:
                name = normalize_name(row.get('氏名', ''))
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row['希望日'],
                    shift_type=row['勤務区分'],
                    priority=row.get('優先度', 1)
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row.get('氏名')}")
            except Exception as e:
                messages.warning(request, f"エラー（{row.get('氏名')}）：{e}")
        messages.success(request, 'CSVインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_csv.html')

