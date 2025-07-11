# ---- Django標準 ----
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.forms import modelformset_factory
from django.core.exceptions import PermissionDenied, ValidationError

# ---- 外部ライブラリ ----
from xhtml2pdf import pisa
from io import BytesIO, StringIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from collections import defaultdict
import calendar

# ---- Python標準ライブラリ ----
import os
import io
import json
import csv
import tempfile
from datetime import date, datetime
from pathlib import Path

# ---- 自作ユーティリティ ----
from .utils import normalize_name

# ---- アプリ内モデル／フォーム ----
from .models import (
    Shift, Employee, ShiftRequest, HolidayRequest,
    LeaveBalance, Committee, OnCallShift, ShiftRequirement,
    DialysisPatient, DialysisStaff
)
from .forms import (
    ShiftForm, ShiftRequestForm, ShiftApprovalForm,
    HolidayRequestForm, ShiftEditForm, DialysisPatientForm, DialysisStaffForm,
    EmployeeForm
)


# ✅ 権限チェック
def staff_required(view_func):
    return user_passes_test(lambda u: u.is_staff)(view_func)

# ✅ メール通知関数
def send_notification_email(subject, message, recipient_email): 
    try:
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [recipient_email])
    except Exception as e:
        print(f"メール送信失敗: {e}")

# 📅 カレンダービュー
@login_required
def calendar_view(request):
    return render(request, 'shifts/calendar_full.html')

@login_required
def calendar_event_api(request):
    events = []
    for shift in Shift.objects.all():
        events.append({
            "title": f"[確定] {shift.employee.name}（{shift.shift_type}）",
            "start": str(shift.shift_date),
            "color": "#007bff"
        })
    for req in ShiftRequest.objects.all():
        events.append({
            "title": f"[希望] {req.employee.name}（{req.shift_type}）",
            "start": str(req.requested_date),
            "color": "#28a745"
        })
    return JsonResponse(events, safe=False)

# ✅ 透析日誌 関連ビュー
@login_required
def dialysis_daily_view(request, selected_date=None):
    selected_date = selected_date or date.today()
    patients = DialysisPatient.objects.filter(date=selected_date)
    staff = DialysisStaff.objects.filter(date=selected_date)
    return render(request, 'dialysis/daily_view.html', {
        'patients': patients, 'staff': staff, 'date': selected_date
    })

@login_required
def dialysis_register_view(request):
    if not request.user.is_staff and datetime.now().day > 17:
        raise PermissionDenied("17日以降は編集できません（スタッフ専用）")

    PatientFormSet = modelformset_factory(DialysisPatient, form=DialysisPatientForm, extra=5)
    StaffFormSet = modelformset_factory(DialysisStaff, form=DialysisStaffForm, extra=5)

    if request.method == 'POST':
        pf = PatientFormSet(request.POST, queryset=DialysisPatient.objects.none())
        sf = StaffFormSet(request.POST, queryset=DialysisStaff.objects.none())
        if pf.is_valid() and sf.is_valid():
            pf.save()
            sf.save()
            return redirect('dialysis_daily')
    else:
        pf = PatientFormSet(queryset=DialysisPatient.objects.none())
        sf = StaffFormSet(queryset=DialysisStaff.objects.none())

    return render(request, 'dialysis/register.html', {'pf': pf, 'sf': sf})

@login_required
def import_dialysis_csv(request):
    if request.method == "POST" and request.FILES.get('file'):
        df = pd.read_excel(request.FILES['file'])
        for _, row in df.iterrows():
            if not pd.isna(row['名前']):
                DialysisPatient.objects.create(
                    name=row['名前'],
                    dialysis_time=row['透析時間'],
                    gender=row['性別'],
                    area=row['地域'],
                    remarks=row.get('備考', ''),
                    date=row.get('日付', date.today())
                )
        messages.success(request, "透析日誌データをインポートしました。")
        return redirect('dialysis_daily')
    return render(request, 'dialysis/import.html')

@login_required
def dialysis_pdf_export(request, date_str):
    target_date = date.fromisoformat(date_str)
    patients = DialysisPatient.objects.filter(date=target_date)
    staff = DialysisStaff.objects.filter(date=target_date)
    template = get_template('dialysis/pdf_template.html')
    html = template.render({'patients': patients, 'staff': staff, 'date': target_date})
    response = BytesIO()
    pisa.CreatePDF(html, dest=response)
    return HttpResponse(response.getvalue(), content_type='application/pdf')

@login_required
def dialysis_calendar_api_events(request):
    events = []
    dates = DialysisPatient.objects.values_list('date', flat=True).distinct()
    for d in dates:
        events.append({'title': '透析あり', 'start': d.isoformat()})
    return JsonResponse(events, safe=False)

# 🧑‍💼 ユーザー登録／ログイン／ログアウト
def signup_view(request):
    form = UserCreationForm(request.POST or None)
    if form.is_valid():
        user = form.save()
        messages.success(request, 'ユーザー登録完了。ログインしてください。')
        admin_emails = [emp.email for emp in Employee.objects.filter(position="管理者")]
        for email in admin_emails:
            send_notification_email(
                subject="【新規ユーザー登録】",
                message=f"新しいユーザー {user.username} さんが登録されました。",
                recipient_email=email
            )
        return redirect('login')
    elif request.method == 'POST':
        messages.error(request, '登録に失敗しました。入力内容をご確認ください。')
    return render(request, 'registration/signup.html', {'form': form})

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def profile(request):
    return render(request, 'shifts/profile.html', {'user': request.user})

@login_required
def homepage(request):
    return render(request, 'shifts/homepage.html')

# 📥 勤務希望（JS API登録＋17日制限）
@csrf_exempt
@login_required
def shift_request_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("date")
            shift_type = data.get("type", "希望")
            requested_date = datetime.strptime(date_str, "%Y-%m-%d").date()

            if requested_date.day > 17 and not request.user.is_staff:
                return JsonResponse({"error": "17日以降の希望登録はできません"}, status=400)

            ShiftRequest.objects.create(
                employee=request.user.employee,
                requested_date=requested_date,
                shift_type=shift_type,
                priority=1
            )
            return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

# 📋 勤務希望一覧／休暇申請
@login_required
def shift_request_list(request):
    requests = ShiftRequest.objects.all().order_by('requested_date')
    return render(request, 'shifts/shift_request_list.html', {'requests': requests})

@login_required
def add_holiday_request(request):
    form = HolidayRequestForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, '休暇申請が完了しました！')
        return redirect('shift_request_list')
    return render(request, 'shifts/add_holiday_request.html', {'form': form})

# 📤 勤務希望出力（Excel / CSV / PDF）
@login_required
def export_shift_requests_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務希望"
    ws.append(["氏名", "希望日", "勤務区分", "優先度"])
    for r in ShiftRequest.objects.all():
        ws.append([r.employee.name, r.requested_date, r.shift_type, r.priority])

    output_path = os.path.join(settings.BASE_DIR, 'shifts', 'temp_exports', 'shift_requests.xlsx')
    wb.save(output_path)

    with open(output_path, 'rb') as f:
        response = HttpResponse(
            f.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = 'attachment; filename="shift_requests.xlsx"'
    return response

@login_required
def export_shift_requests_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="shift_requests.csv"'
    writer = csv.writer(response)
    writer.writerow(['氏名', '希望日', '勤務区分', '優先度'])
    for r in ShiftRequest.objects.all():
        writer.writerow([r.employee.name, r.requested_date, r.shift_type, r.priority])
    return response

@login_required
def export_shift_pdf(request):
    shifts = Shift.objects.order_by('shift_date')
    today = datetime.now()
    html = get_template('shifts/shift_pdf.html').render({'shifts': shifts, 'today': today})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="shift_report.pdf"'
    pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    return response

# 📥 インポート（Excel）
@login_required
def import_shift_requests_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        wb = load_workbook(request.FILES['excel_file'])
        ws = wb.active
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            try:
                name = normalize_name(row[0])
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row[1],
                    shift_type=row[2],
                    priority=row[3] or 1
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row[0]}")
            except Exception as e:
                messages.warning(request, f"エラー（{row[0]}）：{e}")
        messages.success(request, 'Excelインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_excel.html')

# 📥 インポート（CSV）
@login_required
def import_shift_requests_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        df = pd.read_csv(request.FILES['csv_file'])
        for _, row in df.iterrows():
            try:
                name = normalize_name(row.get('氏名', ''))
                emp = Employee.objects.get(name=name)
                ShiftRequest.objects.create(
                    employee=emp,
                    requested_date=row['希望日'],
                    shift_type=row['勤務区分'],
                    priority=row.get('優先度', 1)
                )
            except Employee.DoesNotExist:
                messages.warning(request, f"該当従業員が存在しません: {row.get('氏名')}")
            except Exception as e:
                messages.warning(request, f"エラー（{row.get('氏名')}）：{e}")
        messages.success(request, 'CSVインポート完了！')
        return redirect('shift_request_list')
    return render(request, 'shifts/import_shift_csv.html')

# 🤖 自動割当（ルールベースAI風）
@staff_required
def auto_assign(request):
    assigned = set()
    assigned_counts = defaultdict(lambda: defaultdict(int))
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']:
        week_day_index = list(calendar.day_name).index(day) + 2
        requests = ShiftRequest.objects.filter(requested_date__week_day=week_day_index).order_by('priority')
        requirements = ShiftRequirement.objects.filter(day_of_week=day).first()
        if not requirements:
            continue
        counts = {
            '看護師': requirements.nurse_required,
            '臨床工学技士': requirements.engineer_required,
            '介護福祉士': requirements.assistant_required
        }
        for req in requests:
            emp_id = req.employee.id
            date_key = (emp_id, req.requested_date)
            position = req.employee.position
            if counts.get(position, 0) > 0 and date_key not in assigned:
                Shift.objects.create(
                    employee=req.employee,
                    shift_date=req.requested_date,
                    shift_type=req.shift_type,
                    is_approved=True
                )
                assigned.add(date_key)
                assigned_counts[emp_id][day] += 1
                counts[position] -= 1
    messages.success(request, '最適化された勤務割当が完了しました。')
    return redirect('shift_list')

# 📊 月次集計（ポジション別）
@login_required
def monthly_shift_summary(request):
    summary = Shift.objects.values('employee__position', month=TruncMonth('shift_date')) \
        .annotate(count=Count('id')).order_by('month', 'employee__position')
    return render(request, 'shifts/summary.html', {'summary': summary})

# 📋 管理ダッシュボード（未承認リスト付き）
@staff_required
def admin_dashboard(request):
    shifts = Shift.objects.all().order_by('-shift_date')
    unapproved = ShiftRequest.objects.filter(
        priority__gte=1,
        requested_date__lte=date.today(),
        employee__isnull=False,
        shift_type__isnull=False,
        approved=False
    )
    employees = Employee.objects.all()

    monthly_summary = Shift.objects.values('employee__position', month=TruncMonth('shift_date')) \
        .annotate(count=Count('id')).order_by('month', 'employee__position')

    chart_data = {}
    for entry in monthly_summary:
        month_str = entry['month'].strftime('%Y-%m')
        position = entry['employee__position']
        chart_data.setdefault(month_str, {})
        chart_data[month_str][position] = entry['count']

    context = {
        "shifts": shifts,
        "unapproved": unapproved,
        "employees": employees,
        "total_requests": ShiftRequest.objects.count(),
        "total_shifts": shifts.count(),
        "today": date.today(),
        "chart_data": json.dumps(chart_data)
    }
    return render(request, "shifts/admin_dashboard.html", context)

# 🖥️ 3ブロック勤務表テンプレート（Bootstrap HTML）表示
@login_required
def view_generated_shift_template(request):
    return render(request, 'templates_generated/template_shift_3block_colored.html')

# 🔒 シフト削除確認画面（分離型）
@login_required
def confirm_delete_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    return render(request, 'shifts/confirm_delete_shift.html', {'shift': shift})

# 🧭 テンプレートリンク確認用（開発時用）
@login_required
def template_links_view(request):
    return render(request, 'shifts/template_links.html')

# 📋 全機能ナビゲーション（管理UI）
@login_required
def dashboard_view(request):
    links = [
        {"title": "シフト一覧", "url": reverse("shift_list"), "label": "シフト管理"},
        {"title": "勤務希望一覧", "url": reverse("shift_request_list"), "label": "希望一覧"},
        {"title": "休暇申請", "url": reverse("add_holiday_request"), "label": "休暇申請"},
        {"title": "月次統計", "url": reverse("monthly_shift_summary"), "label": "統計表示"},
        {"title": "PDF出力", "url": reverse("export_shift_pdf"), "label": "シフトPDF"},
        {"title": "カレンダー", "url": reverse("calendar_view"), "label": "フルカレンダー"},
        {"title": "透析登録", "url": reverse("dialysis_register"), "label": "透析入力"},
        {"title": "透析日誌", "url": reverse("dialysis_daily"), "label": "透析日誌"},
        {"title": "テンプレート表示", "url": reverse("shift_template_3block"), "label": "テンプレート表示"},
        {"title": "Django Admin", "url": "/admin/", "label": "管理画面"},
        {"title": "CSV取込", "url": reverse("upload_shifts_csv"), "label": "CSVアップロード"},
    ]
    return render(request, 'dashboard.html', {'links': links})

# 🧾 エラーログCSVダウンロード
@staff_required
def download_error_log_csv(request):
    file_name = request.GET.get('file') or request.session.get('error_log_file')
    if not file_name:
        messages.error(request, "エラーログファイルが見つかりません。")
        return redirect('shift_list')

    file_path = os.path.join(settings.BASE_DIR, 'shifts', 'temp_exports', file_name)
    if not os.path.exists(file_path):
        messages.error(request, "一時ファイルが削除されています。")
        return redirect('shift_list')

    with open(file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response

# ➕ シフト追加
@login_required
def add_shift(request):
    form = ShiftForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, '新しいシフトが追加されました！')
        return redirect('shift_list')
    return render(request, 'shifts/add_shift.html', {'form': form})

# ✏️ シフト編集
@login_required
def edit_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            messages.success(request, 'シフトが更新されました。')
            return redirect('shift_list')
    else:
        form = ShiftForm(instance=shift)
    return render(request, 'shifts/edit_shift.html', {'form': form, 'shift': shift})

# ❌ シフト削除（確認ページ付き）
@login_required
def delete_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        shift.delete()
        messages.success(request, f"シフト {shift.employee.name} - {shift.shift_date} が削除されました。")
        return redirect('shift_list')
    return render(request, 'shifts/confirm_delete_shift.html', {'shift': shift})

# ✅ シフト承認
@login_required
def approve_shift(request, shift_id):
    shift = get_object_or_404(Shift, id=shift_id)
    if request.method == 'POST':
        shift.is_approved = True
        shift.save()
        messages.success(request, f"シフト {shift.employee.name} - {shift.shift_date} が承認されました。")
        return redirect('shift_list')
    return render(request, 'shifts/approve_shift.html', {'shift': shift})

# 📋 シフト一覧（検索・絞り込み付き）
@login_required
def shift_list(request):
    shifts = Shift.objects.all()

    # フィルター機能
    employee_name = request.GET.get("employee")
    date = request.GET.get("date")
    shift_type = request.GET.get("shift_type")

    if employee_name:
        shifts = shifts.filter(employee__name__icontains=employee_name)
    if date:
        shifts = shifts.filter(shift_date=date)
    if shift_type:
        shifts = shifts.filter(shift_type__icontains=shift_type)

    return render(request, 'shifts/shift_list.html', {'shifts': shifts})

from django.views.decorators.csrf import csrf_exempt

@staff_required
@csrf_exempt
def upload_shifts_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        try:
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            Shift.objects.all().delete()

            success_count = 0
            error_logs = []

            for row in reader:
                name = normalize_name(row.get('氏名', ''))
                raw_date = row.get('日付', '').strip()
                shift_type = row.get('勤務区分')
                location = row.get('配置先') or ""
                notes = row.get('備考') or ""

                if not name:
                    error_logs.append({'氏名': name, '日付': raw_date, 'エラー内容': '氏名が空です'})
                    continue

                shift_date = None
                for fmt in ['%Y/%m/%d', '%Y-%m-%d']:
                    try:
                        shift_date = datetime.strptime(raw_date, fmt).date()
                        break
                    except ValueError:
                        continue

                if not shift_date:
                    error_logs.append({'氏名': name, '日付': raw_date, 'エラー内容': '日付形式エラー'})
                    continue

                try:
                    employee = Employee.objects.get(name=name)
                    Shift.objects.create(
                        employee=employee,
                        shift_date=shift_date,
                        shift_type=shift_type,
                        location=location,
                        notes=notes,
                        is_approved=True
                    )
                    success_count += 1
                except Employee.DoesNotExist:
                    error_logs.append({'氏名': name, '日付': raw_date, 'エラー内容': '該当職員がいません'})
                except Exception as e:
                    error_logs.append({'氏名': name, '日付': raw_date, 'エラー内容': str(e)})

            total = success_count + len(error_logs)
            messages.success(request, f"アップロード完了：成功 {success_count} 件 / エラー {len(error_logs)} 件 / 合計 {total} 件")

            if error_logs:
                export_dir = os.path.join(settings.BASE_DIR, 'shifts', 'temp_exports')
                os.makedirs(export_dir, exist_ok=True)
                error_log_path = os.path.join(export_dir, f'error_log_{datetime.now().strftime("%Y%m%d%H%M%S")}.csv')

                with open(error_log_path, 'w', encoding='utf-8-sig', newline='') as csvfile:
                    fieldnames = ['氏名', '日付', 'エラー内容']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    for row in error_logs:
                        writer.writerow(row)

                request.session['error_log_file'] = os.path.basename(error_log_path)
                messages.warning(request, "一部エラーがありました。エラーログをダウンロードしてください。")

            return redirect('shift_list')

        except Exception as e:
            messages.error(request, f"CSV読み込みエラー: {e}")
            return redirect('shift_list')

    return render(request, 'shifts/upload_shifts_csv.html')

@login_required
def employee_list(request):
    employees = Employee.objects.all()
    return render(request, 'shifts/employee_list.html', {'employees': employees})

@login_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'スタッフを登録しました。')
            return redirect('employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'staff/employee_form.html', {'form': form})

@login_required
def upload_employee_csv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded = csv_file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)

        for row in reader:
            try:
                user, _ = User.objects.get_or_create(username=row['user'], defaults={'email': row['email']})
                Employee.objects.create(
                    name=row['name'],
                    name_kana=row.get('name_kana') or '',
                    gender=row.get('gender') or '',
                    birth_date=parse_date(row.get('birth_date')),
                    email=row['email'],
                    phone=row.get('phone') or '',
                    user=user,
                    department=row['department'],
                    assigned_locations=json.loads(row['assigned_locations'] or '{}'),
                    position=row['position'],
                    employment_type=row['employment_type'],
                    hire_date=parse_date(row['hire_date']),
                    work_constraints=json.loads(row.get('work_constraints') or '{}'),
                    qualifications=json.loads(row.get('qualifications') or '[]'),
                    work_details=json.loads(row['work_details']),
                    committees=json.loads(row.get('committees') or '[]'),
                    teams=json.loads(row.get('teams') or '[]'),
                    clinical_ladder=row.get('clinical_ladder') or '',
                    yearly_goals=row.get('yearly_goals') or '',
                    career_history=row.get('career_history') or '',
                    ability_rating=json.loads(row.get('ability_rating') or '{}'),
                    performance_evaluation=json.loads(row.get('performance_evaluation') or '{}'),
                    admin_notes=row.get('admin_notes') or '',
                    interview_history=json.loads(row.get('interview_history') or '[]'),
                    notes=row.get('notes') or '',
                    is_active=row.get('is_active', 'TRUE').upper() == 'TRUE'
                )
            except Exception as e:
                messages.error(request, f"エラー: {e}")
                continue

        messages.success(request, 'CSVからスタッフ情報をインポートしました。')
        return redirect('employee_list')

    return render(request, 'staff/upload_employee_csv.html')


@login_required
def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, 'スタッフ情報を更新しました。')
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'shifts/employee_edit.html', {'form': form, 'employee': employee})

@login_required
def employee_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    return render(request, 'shifts/employee_detail.html', {'employee': employee})

@login_required
def delete_duplicates(request):
    seen = set()
    duplicates = []
    for emp in Employee.objects.all():
        if emp.name in seen:
            duplicates.append(emp)
        else:
            seen.add(emp.name)
    for dup in duplicates:
        dup.delete()
    messages.success(request, f"{len(duplicates)} 件の重複従業員を削除しました。")
    return redirect('employee_list')

@user_passes_test(lambda u: u.is_staff)
def upload_fixture(request):
    if request.method == 'POST' and request.FILES.get('data_file'):
        data_file = request.FILES['data_file']
        ext = data_file.name.split('.')[-1].lower()

        try:
            if ext == 'json':
                data = json.load(data_file)
                created_count = 0
                for entry in data:
                    fields = entry.get("fields", entry)
                    obj, created = Employee.objects.update_or_create(
                        name=fields.get("name"),
                        defaults=fields
                    )
                    if created:
                        created_count += 1
                messages.success(request, f"✅ JSONから {created_count} 件登録しました。")
            elif ext == 'csv':
                decoded_file = data_file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                created_count = 0
                for row in reader:
                    # CSV内のJSON文字列項目を読み取り
                    for k in ['assigned_locations', 'work_constraints', 'qualifications',
                              'work_details', 'committees', 'teams', 'ability_rating',
                              'performance_evaluation', 'interview_history']:
                        if row.get(k):
                            try:
                                row[k] = json.loads(row[k])
                            except Exception:
                                row[k] = {}

                    row['hire_date'] = parse_date(row.get('hire_date')) if row.get('hire_date') else None
                    row['birth_date'] = parse_date(row.get('birth_date')) if row.get('birth_date') else None
                    row['is_active'] = row.get('is_active', 'TRUE').upper() == 'TRUE'

                    obj, created = Employee.objects.update_or_create(
                        name=row.get("name"),
                        defaults=row
                    )
                    if created:
                        created_count += 1

                messages.success(request, f"✅ CSVから {created_count} 件登録しました。")
            else:
                messages.error(request, "対応していないファイル形式です（.json または .csv のみ）")
        except Exception as e:
            messages.error(request, f"エラーが発生しました: {str(e)}")

        return redirect('employee_list')

    return render(request, 'shifts/upload_fixture.html')